"""
키워드 매칭 자동화 - 문자/단어/SBERT + LLM 검증/매칭 버전

[매칭 순서]
1단계 : 정확 일치 (100%)
2단계 : 문자/단어 SequenceMatcher ≥90%
3단계 : 일본어 브랜드 정규화 후 문자 유사도 ≥70% (일본어 전용)
4단계 : SBERT 의미 유사도 ≥80%
         └ ≥90%        → 확정
         └ 80~89%      → LLM 검증 (같은 의미면 유지 / 다르면 미매칭 전환)
5단계 : 미매칭 BL + 반환된 KL → SBERT 후보 선별(≥50%) → LLM 최종 매칭
6단계 : 임계값 무시 최종 패스 → SBERT 상위K 무조건 LLM 전송

[색상]
흰색=100%, 노랑=99%+, 주황=95~98%, 빨강=90~94%, 회색=없음
연보라=SBERT 확정, 연분홍=검토필요, 연초록=LLM 매칭, 연파랑=LLM 검증통과, 연노랑=LLM 최종
"""

import os
import re
import time
from difflib import SequenceMatcher
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ============================================================
# 설정
# ============================================================

INPUT_FILE   = r"C:\Users\yunji\OneDrive\문서\통합 문서1_20260315.xlsx"
OUTPUT_FILE  = r"C:\Users\yunji\OneDrive\문서\통합 문서1_매칭결과_20260405.xlsx"

S_CORE_FILE        = r"C:\Users\yunji\Documents\Samsung\맵핑 검수\S_core_initial_processing_v1.xlsx"
S_CORE_OUTPUT_FILE = r"C:\Users\yunji\Documents\Samsung\맵핑 검수\S_core_updated.xlsx"
S_CORE_SHEET       = "Sheet1"
S_CORE_DATE_TARGET = "2026-03-22"

OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")

BL_SHEET      = "브랜드라이트"
BL_KW_COL     = 1
BL_LANG_COL   = 2
BL_RESULT_COL = 3
BL_HEADER_ROW = 1

REF_SHEET      = "기존 키워드리스트"
REF_HEADER_ROW = 1

PHASES = [("char", 90), ("word", 90)]

# Step 4: SBERT 매칭 임계값
SBERT_MODEL              = "paraphrase-multilingual-MiniLM-L12-v2"
SBERT_MATCH_THRESHOLD    = 75   # SBERT 매칭 최소 임계값
SBERT_MATCH_THRESHOLD_JA = 60   # 일본어 전용
SBERT_CONFIRM_THRESHOLD  = 90   # 이 이상은 LLM 검증 없이 바로 확정
SBERT_AUTO_PASS          = 80
SBERT_CHAR_THRESHOLD     = 70

# Step 5: LLM 매칭용 SBERT 후보 선별 임계값
SBERT_CAND_THRESHOLD     = 50
SBERT_CAND_THRESHOLD_JA  = 40
SBERT_CAND_THRESHOLD_KO  = 55
SBERT_KO_SUFFIX_MIN      = 40   # Step 5 LLM 후보 필터 (한국어: suffix 유사도 미달 시 후보 제외)
SBERT_KO_SUFFIX_CONFIRM  = 60   # 이 이상이어야 Step 4에서 SBERT 확정 (미만이면 LLM 검증)
SBERT_KO_GROUPS          = {"mx_ko", "co.a_ko"}

# LLM 설정
LLM_MODEL      = "gpt-4o-mini"
LLM_CANDIDATES = 5
LLM_BATCH_SIZE = 20
LLM_PARALLEL   = 10

KL_LANG_COLS = {
    "MX_en":    [1],      "MX_ko":   [2],      "MX_pt":   [3],      "MX_it":  [4],
    "MX_de":    [5],      "MX_ja":   [6],      "MX_id":   [7],      "MX_es":  [8],
    "MX_fr":    [9],      "Co.A_en": [10],      "Co.A_ko": [11],      "Co.A_pt": [12],
    "Co.A_it":  [13],      "Co.A_de": [14],      "Co.A_ja": [15],      "Co.A_id": [16],
    "Co.A_es":  [17],      "Co.A_fr": [18]


}
_KL_COL_LANG = {col - 1: grp.lower()
                for grp, cols in KL_LANG_COLS.items()
                for col in cols}

# ============================================================

def normalize(text: str) -> str:
    return re.sub(r"[^\w\s]", "", str(text).lower().strip())


JA_BRAND_NORM = {
    "macbook": "マックブック", "iphone": "アイフォン", "ipad": "アイパッド",
    "ipod": "アイポッド", "samsung": "サムスン", "galaxy": "ギャラクシー",
    "apple": "アップル", "ultra": "ウルトラ", "mac": "マック",
}

# 브랜드명 외 일반 카타카나 기술 용어 (normalize_ja 변환용, ja_brand_residual 제거 대상 아님)
JA_KATA_NORM = {
    "smartphone":    "スマートフォン",
    "camera":        "カメラ",
    "gaming":        "ゲーミング",
    "game":          "ゲーム",
    "display":       "ディスプレイ",
    "battery":       "バッテリー",
    "memory":        "メモリ",
    "storage":       "ストレージ",
    "performance":   "パフォーマンス",
    "system":        "システム",
    "color":         "カラー",
    "variation":     "バリエーション",
    "high end":      "ハイエンド",
    "photo":         "フォト",
    "editor":        "エディター",
    "chat":          "チャット",
    "assist":        "アシスト",
    "reviewer":      "レビュアー",
    "review":        "レビュー",
    "full":          "フル",
    "plus":          "プラス",
    "pen":           "ペン",
    "spec":          "スペック",
    "charge":        "チャージ",
    "update":        "アップデート",
    "application":   "アプリケーション",
    "app":           "アプリ",
    "screen":        "スクリーン",
    "touch":         "タッチ",
    "processor":     "プロセッサ",
    "sensor":        "センサー",
    "feature":       "フィーチャー",
    "mode":          "モード",
    "internet":      "インターネット",
    "network":       "ネットワーク",
    "wireless":      "ワイヤレス",
    "bluetooth":     "ブルートゥース",
    "video":         "ビデオ",
    "content":       "コンテンツ",
    "support":       "サポート",
    "service":       "サービス",
    "design":        "デザイン",
    "power":         "パワー",
    "zoom":          "ズーム",
    "speaker":       "スピーカー",
}

_JA_KATA_TO_ASCII  = {v: k for k, v in JA_BRAND_NORM.items()}
_JA_KATA_TO_ASCII.update({v: k for k, v in JA_KATA_NORM.items()})
_JA_BRAND_TOKENS   = sorted(JA_BRAND_NORM.keys(), key=len, reverse=True)


def normalize_ja(text: str) -> str:
    t = str(text).lower().strip()
    for kata, ascii_w in sorted(_JA_KATA_TO_ASCII.items(), key=lambda x: -len(x[0])):
        t = t.replace(kata, f" {ascii_w} ")
    return normalize(re.sub(r'\s+', ' ', t).strip())


def ja_brand_residual(text: str) -> str:
    t = normalize_ja(text)
    for brand in _JA_BRAND_TOKENS:
        t = t.replace(brand, " ")
    return re.sub(r'\s+', ' ', t).strip()


def strip_common_prefix(a: str, b: str) -> tuple:
    ta, tb = a.split(), b.split()
    i = 0
    while i < len(ta) and i < len(tb) and ta[i] == tb[i]:
        i += 1
    return " ".join(ta[i:]), " ".join(tb[i:])


def suffix_score(a: str, b: str) -> float:
    sa, sb = strip_common_prefix(a, b)
    if not sa and not sb: return 100.0
    if not sa or not sb:  return 0.0
    return char_score(sa, sb)


# 한국어 질문 어미 (길이 순 정렬 — 긴 것부터 매칭)
_KO_ENDINGS = sorted([
    "을 수 있나요", "을 수 있어요", "을 수 있을까요",
    "수 있나요", "수 있어요", "수 있을까요",
    "인가요", "이나요", "있나요", "없나요", "하나요", "되나요",
    "는가요", "을까요", "할까요", "일까요",
    "나요", "어요", "아요", "네요", "죠",
], key=len, reverse=True)


def _strip_ko_ending(text: str) -> str:
    for ending in _KO_ENDINGS:
        if text.endswith(ending):
            return text[:-len(ending)].strip()
    return text


def ko_suffix_score(a: str, b: str) -> float:
    """한국어 전용 suffix 비교 — 질문 어미 제거 후 핵심 내용만 비교"""
    sa, sb = strip_common_prefix(a, b)
    if not sa and not sb: return 100.0
    if not sa or not sb:  return 0.0
    return char_score(_strip_ko_ending(sa), _strip_ko_ending(sb))


def char_score(a: str, b: str) -> float:
    return round(SequenceMatcher(None, a, b).ratio() * 100, 1)


def tokenize(text: str) -> list:
    parts = re.findall(r'[a-z0-9]+|[^\x00-\x7f\s]+', text)
    if not parts:
        return list(text)
    tokens = []
    for p in parts:
        if re.search(r'[^\x00-\x7f]', p) and len(p) > 2:
            tokens.extend(p[i:i + 2] for i in range(len(p) - 1))
        else:
            tokens.append(p)
    return tokens if tokens else list(text)


def word_score(a: str, b: str) -> float:
    ta, tb = tokenize(a), tokenize(b)
    if not ta or not tb: return 0.0
    return round(SequenceMatcher(None, ta, tb).ratio() * 100, 1)


def load_ref(path: str, sheet: str, header_row: int) -> dict:
    wb  = load_workbook(path, read_only=True, data_only=True)
    ws  = wb[sheet]
    ref = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row:
            continue
        for ci, val in enumerate(row):
            if ci not in _KL_COL_LANG or val is None or not str(val).strip():
                continue
            original = str(val).strip()
            grp      = _KL_COL_LANG[ci]
            norm     = normalize(original)
            if norm:
                ref[(norm, grp)] = original
    wb.close()
    return ref


def row_fill(score: float) -> PatternFill:
    if   score == 100: color = "FFFFFF"
    elif score >= 99:  color = "FFFF99"
    elif score >= 95:  color = "FFD966"
    elif score >= 90:  color = "FF9999"
    else:              color = "F2F2F2"
    return PatternFill("solid", fgColor=color)


def calc_pairs(ref_pool, bl_pool, bl_lang, mode="char", min_score=0):
    score_fn  = char_score if mode == "char" else word_score
    bl_by_grp = defaultdict(list)
    for bl_norm in bl_pool:
        bl_by_grp[bl_lang.get(bl_norm, "")].append(bl_norm)

    pairs = []
    for (ref_norm, grp) in ref_pool:
        candidates = bl_by_grp.get(grp)
        if not candidates:
            continue
        ref_len = len(ref_norm)
        for bl_norm in candidates:
            bl_txt = bl_norm[0]          # (kw_norm, lang_grp) 튜플에서 텍스트만
            bl_len = len(bl_txt)
            if bl_len == 0 or ref_len / bl_len > 2 or bl_len / ref_len > 2:
                continue
            if mode == "char" and min_score > 0:
                sm = SequenceMatcher(None, ref_norm, bl_txt)
                if sm.real_quick_ratio() * 100 < min_score: continue
                if sm.quick_ratio()      * 100 < min_score: continue
            score = score_fn(ref_norm, bl_txt)
            if score >= min_score:
                pairs.append(((ref_norm, grp), bl_norm, score))

    pairs.sort(key=lambda x: x[2], reverse=True)
    return pairs


def ja_brand_pairs(ref_pool, bl_pool, bl_lang, min_score=70):
    # 그룹별로 분류 (mx_ja / co.a_ja 섞이지 않도록)
    ja_bl_by_grp = defaultdict(list)
    for bn in bl_pool:
        grp = bl_lang.get(bn, "")
        if grp.endswith("_ja"):
            ja_bl_by_grp[grp].append((bn, ja_brand_residual(bn[0])))

    if not ja_bl_by_grp:
        return []

    pairs = []
    for (ref_norm, grp) in ref_pool:
        if not grp.endswith("_ja"):
            continue
        ja_bl = ja_bl_by_grp.get(grp, [])  # 같은 그룹 BL만
        if not ja_bl:
            continue
        ref_res    = ja_brand_residual(ref_norm)
        best_score = 0
        best_bl    = ""
        for bn, bn_res in ja_bl:
            if not ref_res and not bn_res:
                best_score = 100.0; best_bl = bn; break
            if not ref_res or not bn_res:
                continue
            ref_len = len(ref_res); bl_len = len(bn_res)
            if ref_len / max(bl_len, 1) > 2 or bl_len / max(ref_len, 1) > 2:
                continue
            s = char_score(ref_res, bn_res)
            if s > best_score:
                best_score = s; best_bl = bn
        if best_bl and best_score >= min_score:
            pairs.append(((ref_norm, grp), best_bl, best_score))

    pairs.sort(key=lambda x: x[2], reverse=True)
    return pairs


def sbert_match_all(ref_pool, bl_pool, bl_lang, bl_orig, model=None):
    """
    SBERT로 SBERT_MATCH_THRESHOLD 이상 모든 쌍 반환 + 모델 반환
    (≥90% / 80-89% 분리는 run()에서 처리)
    """
    import numpy as np
    from sentence_transformers import SentenceTransformer

    if model is None:
        print(f"  SBERT 모델 로드 중... ({SBERT_MODEL})")
        model = SentenceTransformer(SBERT_MODEL)

    bl_by_grp  = defaultdict(list)
    for bl_norm in bl_pool:
        bl_by_grp[bl_lang.get(bl_norm, "")].append(bl_norm)

    ref_by_grp = defaultdict(list)
    for ref_key in ref_pool:
        ref_by_grp[ref_key[1]].append(ref_key)

    all_pairs = []

    for grp, ref_keys in ref_by_grp.items():
        bl_norms = bl_by_grp.get(grp, [])
        if not bl_norms:
            continue

        is_ja   = grp.endswith("_ja")
        grp_thr = SBERT_MATCH_THRESHOLD_JA if is_ja else SBERT_MATCH_THRESHOLD

        if is_ja:
            ref_texts = [ja_brand_residual(ref_pool[rk]) or ref_pool[rk] for rk in ref_keys]
            bl_texts  = [ja_brand_residual(bl_orig.get(bn, bn[0])) or bl_orig.get(bn, bn[0])
                         for bn in bl_norms]
        else:
            ref_texts = [ref_pool[rk] for rk in ref_keys]
            bl_texts  = [bl_orig.get(bn, bn[0]) for bn in bl_norms]

        ref_embs = model.encode(ref_texts, convert_to_numpy=True,
                                show_progress_bar=False, batch_size=64)
        bl_embs  = model.encode(bl_texts,  convert_to_numpy=True,
                                show_progress_bar=False, batch_size=64)

        ref_n = ref_embs / np.linalg.norm(ref_embs, axis=1, keepdims=True)
        bl_n  = bl_embs  / np.linalg.norm(bl_embs,  axis=1, keepdims=True)
        sim   = (ref_n @ bl_n.T) * 100

        is_ko = grp in SBERT_KO_GROUPS
        for ri, ref_key in enumerate(ref_keys):
            for bi in range(len(bl_norms)):
                score = float(sim[ri, bi])
                if score >= grp_thr:
                    if score >= SBERT_AUTO_PASS:
                        # 한국어: suffix가 다르면 LLM 검증으로
                        if is_ko and ko_suffix_score(ref_key[0], bl_norms[bi][0]) < SBERT_KO_SUFFIX_CONFIRM:
                            label = "검토필요"
                        else:
                            label = "SBERT"
                    else:
                        c = max(char_score(ref_key[0], bl_norms[bi][0]),
                                word_score(ref_key[0], bl_norms[bi][0]))
                        label = "SBERT" if c > SBERT_CHAR_THRESHOLD else "검토필요"
                    all_pairs.append((ref_key, bl_norms[bi], round(score, 1), label))

    all_pairs.sort(key=lambda x: x[2], reverse=True)

    # greedy 1:1
    used_ref = set(); used_bl = set(); result = []
    for ref_key, bl_norm, score, label in all_pairs:
        if ref_key in used_ref or bl_norm in used_bl:
            continue
        used_ref.add(ref_key); used_bl.add(bl_norm)
        result.append((ref_key, bl_norm, score, label))

    return result, model


def _llm_call_with_retry(fn, *args, max_retries=6, **kwargs):
    """429 rate-limit 에러 시 지수 백오프로 재시도"""
    delay = 2.0
    for attempt in range(max_retries):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            msg = str(e)
            if "429" in msg or "rate_limit" in msg.lower():
                wait = delay * (2 ** attempt)
                print(f"  429 rate limit — {wait:.0f}초 후 재시도 ({attempt+1}/{max_retries})")
                time.sleep(wait)
            else:
                raise
    return fn(*args, **kwargs)


def ask_llm_batch(client, queries: list) -> list:
    """
    queries: [(bl_text, [kl_candidate_texts, ...]), ...]
    returns: [matched_kl_text or None, ...]
    """
    import json
    if not queries:
        return []

    blocks = []
    for i, (bl_text, candidates) in enumerate(queries):
        cand_lines = "\n".join(f"  {chr(65+j)}. {c}" for j, c in enumerate(candidates))
        blocks.append(f"[{i}] BL: {bl_text}\nKL candidates:\n{cand_lines}")

    prompt = (
        "You are a multilingual keyword matching assistant.\n"
        "For each BL keyword below, decide which KL candidate (if any) has the SAME SEARCH INTENT — "
        "meaning a person searching that BL keyword would want the exact same information or result "
        "as someone searching the matched KL keyword.\n"
        "Different phrasings or translations of the same query = match. "
        "Different products, topics, or intents = no match.\n\n"
        "Return a JSON object: {\"matches\": [{\"idx\": 0, \"match\": \"exact KL text or null\"}, ...]}\n\n"
        "Items:\n\n" + "\n\n".join(blocks)
    )
    try:
        resp   = _llm_call_with_retry(
            client.chat.completions.create,
            model=LLM_MODEL, messages=[{"role": "user", "content": prompt}],
            temperature=0, response_format={"type": "json_object"})
        parsed = json.loads(resp.choices[0].message.content)
        arr    = parsed.get("matches", [])
        results = [None] * len(queries)
        for item in arr:
            idx = item.get("idx"); match = item.get("match")
            if idx is not None and 0 <= idx < len(queries):
                results[idx] = match if match else None
        return results
    except Exception as e:
        print(f"  LLM 응답 오류: {e}")
        return [None] * len(queries)


def ask_llm_verify_batch(client, pairs: list) -> list:
    """
    pairs: [(bl_text, kl_text), ...]
    returns: [True/False, ...]  — 같은 검색 의도면 True
    """
    import json
    if not pairs:
        return []

    blocks = []
    for i, (bl_text, kl_text) in enumerate(pairs):
        blocks.append(f"[{i}] BL: {bl_text}\n    KL: {kl_text}")

    prompt = (
        "You are a multilingual keyword matching verifier.\n"
        "For each BL/KL pair, decide if they have the SAME SEARCH INTENT — "
        "a person searching BL would want the exact same information as someone searching KL.\n"
        "NOTE: Different sentence structure or phrasing is NOT a reason to say false. "
        "Judge only whether the specific information sought is the same.\n\n"
        "SAME intent (true) — different phrasing/structure, but same answer needed:\n"
        "  'Is S26 waterproof?' / 'What is the water resistance rating of S26?' → true\n"
        "  'Best phone for gaming?' / 'Which smartphone has the highest gaming performance?' → true\n"
        "  'Qual è ora la sintesi su S26?' / 'Cos\\'è Now Brief su S26?' → true\n"
        "  'Welches Smartphone eignet sich am besten zum Spielen?' / 'Welches Galaxy-Modell bietet die beste Spieleleistung?' → true\n"
        "  'So entsperren Sie ein eingefrorenes Samsung Galaxy-Handy' / 'Wie kann ich ein Samsung Galaxy-Smartphone wieder entsperren?' → true\n\n"
        "DIFFERENT intent (false) — same topic but the specific question asked is different:\n"
        "  'Is S26 waterproof?' / 'Is S26 good for selfies?' → false\n"
        "  'When can I pre-order S26?' / 'Where can I buy S26?' → false\n"
        "  'Was ist das Besondere am S26/S26+?' / 'Was ist Now Brief auf S26/S26+?' → false\n"
        "    (general overview of the phone ≠ asking about one specific feature)\n"
        "  '¿Cómo mejora la IA el efecto de cámara lenta en el Galaxy S26?' / "
        "'¿Hay ajustes de velocidad o sincronización en las funciones de cámara lenta?' → false\n"
        "    (how AI improves slow-motion ≠ are there speed/sync settings)\n"
        "  'Comment l\\'IA fonctionne-t-elle dans les derniers smartphones?' / "
        "'Comment fonctionnent les générateurs d\\'images basés sur l\\'IA?' → false\n"
        "    (how AI works in smartphones generally ≠ how AI image generators work specifically)\n\n"
        "Return JSON: {\"results\": [{\"idx\": 0, \"same\": true}, {\"idx\": 1, \"same\": false}, ...]}\n\n"
        "Pairs:\n\n" + "\n\n".join(blocks)
    )
    try:
        resp   = _llm_call_with_retry(
            client.chat.completions.create,
            model=LLM_MODEL, messages=[{"role": "user", "content": prompt}],
            temperature=0, response_format={"type": "json_object"})
        parsed = json.loads(resp.choices[0].message.content)
        arr    = parsed.get("results", [])
        results = [False] * len(pairs)
        for item in arr:
            idx = item.get("idx")
            if idx is not None and 0 <= idx < len(pairs):
                results[idx] = bool(item.get("same", False))
        return results
    except Exception as e:
        print(f"  LLM 검증 오류: {e}")
        return [False] * len(pairs)


def ask_llm_verify_strict_batch(client, pairs: list) -> list:
    """
    최종 검수 전용 보수적 검증 — 이미 확정된 쌍을 재검토
    '확실히 다를 때만 False, 조금이라도 같은 의미면 True'
    pairs: [(bl_text, kl_text), ...]
    returns: [True/False, ...]
    """
    import json
    if not pairs:
        return []

    blocks = []
    for i, (bl_text, kl_text) in enumerate(pairs):
        blocks.append(f"[{i}] BL: {bl_text}\n    KL: {kl_text}")

    prompt = (
        "You are a multilingual keyword matching verifier.\n"
        "These pairs were already matched. Re-check if each pair has the SAME SEARCH INTENT.\n\n"
        "SAME intent (say true):\n"
        "  BL: 'ゲーミング性能が最も高いスマートフォン' / KL: 'ゲームに最適なスマートフォンはどれ' → true\n"
        "  BL: 'Qual è ora la sintesi su S26?' / KL: 'Cos'è Now Brief su S26?' → true\n"
        "  BL: 'Is S26 waterproof?' / KL: 'What is the water resistance rating of S26?' → true\n"
        "  BL: '高性能プロセッサを搭載したスマートフォンの中で、マルチタスクとゲーム性能が最も優れているのはどれか'"
        " / KL: '高性能プロセッサを搭載したスマートフォンの中で、マルチタスクとゲーム性能が最も優れている機種はどれですか' → true\n\n"
        "DIFFERENT intent (say false — only when CLEARLY different):\n"
        "  BL: 'ゲーミング性能が最も高いスマートフォン' / KL: '動き回る子供を撮影する際の画像安定化性能' → false\n"
        "  BL: 'Is S26 waterproof?' / KL: 'Is S26 good for selfies?' → false\n"
        "  BL: 'When can I pre-order S26?' / KL: 'Where can I buy S26?' → false\n\n"
        "IMPORTANT: Default to true when in doubt. "
        "Only return false if intents are CLEARLY and OBVIOUSLY different.\n\n"
        "Return JSON: {\"results\": [{\"idx\": 0, \"same\": true}, ...]}\n\n"
        "Pairs:\n\n" + "\n\n".join(blocks)
    )
    try:
        resp   = _llm_call_with_retry(
            client.chat.completions.create,
            model=LLM_MODEL, messages=[{"role": "user", "content": prompt}],
            temperature=0, response_format={"type": "json_object"})
        parsed = json.loads(resp.choices[0].message.content)
        arr    = parsed.get("results", [])
        results = [True] * len(pairs)   # 기본값 True (의심스러우면 유지)
        for item in arr:
            idx = item.get("idx")
            if idx is not None and 0 <= idx < len(pairs):
                results[idx] = bool(item.get("same", True))
        return results
    except Exception as e:
        print(f"  LLM strict 검증 오류: {e}")
        return [True] * len(pairs)   # 오류 시에도 True (매칭 유지)


def llm_verify_pairs(client, to_verify: dict, bl_orig: dict, bl_lang: dict):
    """
    80~89% 매칭된 쌍을 LLM으로 검증 (그룹별 배치)
    to_verify: {bl_norm: (ref_key, kl_text, score, label)}
    returns: (confirmed, rejected)
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    by_grp = defaultdict(list)
    for bl_norm, val in to_verify.items():
        grp = bl_lang.get(bl_norm, "")
        by_grp[grp].append((bl_norm, val))

    batches = []
    for grp, items in by_grp.items():
        for start in range(0, len(items), LLM_BATCH_SIZE):
            batch = items[start:start + LLM_BATCH_SIZE]
            pairs = [(bl_orig.get(bl_norm, bl_norm), kl_text)
                     for bl_norm, (_, kl_text, _, _) in batch]
            batches.append((batch, pairs))

    total_batches = len(batches)
    results_map   = {}
    completed     = 0

    with ThreadPoolExecutor(max_workers=LLM_PARALLEL) as executor:
        futures = {executor.submit(ask_llm_verify_batch, client, pairs): batch
                   for batch, pairs in batches}
        for future in as_completed(futures):
            batch    = futures[future]
            verdicts = future.result()
            completed += 1
            print(f"  LLM 검증 배치 {completed}/{total_batches} 완료")
            for (bl_norm, _), verdict in zip(batch, verdicts):
                results_map[bl_norm] = verdict

    confirmed = {}
    rejected  = {}
    for bl_norm, (ref_key, kl_text, score, _) in to_verify.items():
        if results_map.get(bl_norm, False):
            confirmed[bl_norm] = (ref_key, kl_text, score, "LLM검증")
        else:
            rejected[bl_norm]  = (ref_key, kl_text)

    return confirmed, rejected


def llm_match(ref_pool, bl_pool, bl_lang, bl_orig, sbert_model, client,
              use_threshold=True):
    """
    SBERT 후보 선별 → LLM 최종 매칭
    use_threshold=True : SBERT_CAND_THRESHOLD 이상만 후보
    use_threshold=False: 임계값 무시, SBERT 상위 LLM_CANDIDATES개 무조건 전송 (Step 6)
    """
    import numpy as np
    from concurrent.futures import ThreadPoolExecutor, as_completed

    bl_by_grp  = defaultdict(list)
    for bl_norm in bl_pool:
        bl_by_grp[bl_lang.get(bl_norm, "")].append(bl_norm)

    ref_by_grp = defaultdict(list)
    for ref_key in ref_pool:
        ref_by_grp[ref_key[1]].append(ref_key)

    all_candidates = []

    for grp, ref_keys in ref_by_grp.items():
        bl_norms = bl_by_grp.get(grp, [])
        if not bl_norms:
            continue

        is_ja  = grp.endswith("_ja")
        is_ko  = grp in SBERT_KO_GROUPS
        grp_thr = (SBERT_CAND_THRESHOLD_JA if is_ja else
                   SBERT_CAND_THRESHOLD_KO if is_ko else SBERT_CAND_THRESHOLD)

        if is_ja:
            ref_texts = [ja_brand_residual(ref_pool[rk]) or ref_pool[rk] for rk in ref_keys]
            bl_texts  = [ja_brand_residual(bl_orig.get(bn, bn[0])) or bl_orig.get(bn, bn[0])
                         for bn in bl_norms]
        else:
            ref_texts = [ref_pool[rk] for rk in ref_keys]
            bl_texts  = [bl_orig.get(bn, bn[0]) for bn in bl_norms]

        ref_embs = sbert_model.encode(ref_texts, convert_to_numpy=True,
                                      show_progress_bar=False, batch_size=64)
        bl_embs  = sbert_model.encode(bl_texts,  convert_to_numpy=True,
                                      show_progress_bar=False, batch_size=64)

        ref_n = ref_embs / np.linalg.norm(ref_embs, axis=1, keepdims=True)
        bl_n  = bl_embs  / np.linalg.norm(bl_embs,  axis=1, keepdims=True)
        sim   = (ref_n @ bl_n.T) * 100

        for bi, bl_norm in enumerate(bl_norms):
            candidates = []
            for ri, ref_key in enumerate(ref_keys):
                score = float(sim[ri, bi])
                if use_threshold:
                    if score < grp_thr:
                        continue
                    if is_ko and ko_suffix_score(ref_key[0], bl_norm[0]) < SBERT_KO_SUFFIX_MIN:
                        continue
                candidates.append((score, ref_key))
            if not candidates:
                continue
            candidates.sort(reverse=True)
            all_candidates.append((bl_norm, candidates[:LLM_CANDIDATES], grp))

    if not all_candidates:
        return []

    print(f"  LLM 매칭 대상: {len(all_candidates):,}개 BL  "
          f"(배치 크기: {LLM_BATCH_SIZE}, 병렬 {LLM_PARALLEL}개)")

    kl_orig = {rk: ref_pool[rk] for rk in ref_pool}
    by_grp  = defaultdict(list)
    for item in all_candidates:
        by_grp[item[2]].append(item)

    batches = []
    for grp, grp_items in by_grp.items():
        for start in range(0, len(grp_items), LLM_BATCH_SIZE):
            batch   = grp_items[start:start + LLM_BATCH_SIZE]
            queries = [(bl_orig.get(bl_norm, bl_norm[0]),
                        [kl_orig[rk] for _, rk in cands])
                       for bl_norm, cands, _ in batch]
            batches.append((batch, queries))

    total_batches = len(batches)
    all_pairs     = []
    completed     = 0

    with ThreadPoolExecutor(max_workers=LLM_PARALLEL) as executor:
        futures = {executor.submit(ask_llm_batch, client, queries): (batch, queries)
                   for batch, queries in batches}
        for future in as_completed(futures):
            batch, queries = futures[future]
            matched_texts  = future.result()
            completed += 1
            print(f"  LLM 매칭 배치 {completed}/{total_batches} 완료")
            for (bl_norm, cands, grp), matched_text in zip(batch, matched_texts):
                if not matched_text:
                    continue
                for score, ref_key in cands:
                    if kl_orig.get(ref_key, "") == matched_text:
                        all_pairs.append((ref_key, bl_norm, round(score, 1), "LLM매칭"))
                        break

    all_pairs.sort(key=lambda x: x[2], reverse=True)
    used_ref = set(); used_bl = set(); result = []
    for ref_key, bl_norm, score, label in all_pairs:
        if ref_key in used_ref or bl_norm in used_bl:
            continue
        used_ref.add(ref_key); used_bl.add(bl_norm)
        result.append((ref_key, bl_norm, score, label))
    return result


def llm_match_direct(ref_pool, bl_pool, bl_lang, bl_orig, client):
    """
    Step 6 전용: SBERT 없이 남은 KL 전체를 그룹별로 LLM에 직접 전송
    - Step 5까지 살아남은 키워드만 대상이므로 그룹당 수가 적음
    - BL은 LLM_BATCH_SIZE 단위로 분할, 같은 그룹 KL 전체를 후보로 전송
    """
    import json
    from concurrent.futures import ThreadPoolExecutor, as_completed

    bl_by_grp = defaultdict(list)
    for bl_norm in bl_pool:
        bl_by_grp[bl_lang.get(bl_norm, "")].append(bl_norm)

    ref_by_grp = defaultdict(list)
    for ref_key in ref_pool:
        ref_by_grp[ref_key[1]].append(ref_key)

    # 그룹별로 묶기 (그룹 내 배치는 순차 처리 → greedy 충돌 방지)
    grp_batches = defaultdict(list)
    for grp, bl_norms in bl_by_grp.items():
        ref_keys = ref_by_grp.get(grp, [])
        if not ref_keys:
            continue
        for start in range(0, len(bl_norms), LLM_BATCH_SIZE):
            grp_batches[grp].append((bl_norms[start:start + LLM_BATCH_SIZE], list(ref_keys)))

    if not grp_batches:
        return []

    total_batches = sum(len(v) for v in grp_batches.values())
    print(f"  총 배치: {total_batches:,}개  (그룹 {len(grp_batches)}개 병렬, 그룹 내 순차)")

    def process_batch(grp, batch_bl, ref_keys):
        import json
        kl_orig_local = {rk: ref_pool[rk] for rk in ref_keys}
        kl_texts = [kl_orig_local[rk] for rk in ref_keys]
        bl_texts = [bl_orig.get(bn, bn[0]) for bn in batch_bl]
        bl_lines = "\n".join(f"  {i}. {t}" for i, t in enumerate(bl_texts))
        kl_lines = "\n".join(f"  {i}. {t}" for i, t in enumerate(kl_texts))
        prompt = (
            "You are a multilingual keyword matching assistant.\n"
            "For each BL keyword, find the KL keyword where a person searching BL would want the SAME answer as someone searching KL.\n\n"
            "MATCH — same question, different phrasing or language:\n"
            "  'Can I make stickers on S26?' vs 'Posso creare adesivi su S26?' → MATCH (EN vs IT, same question)\n"
            "  'Is iPhone 17 waterproof?' vs 'What is the water resistance of iPhone 17?' → MATCH (different phrasing, same aspect)\n"
            "  'Best phone for night photos?' vs 'Which phone takes the best low-light shots?' → MATCH (same intent)\n"
            "  'Qual è ora la sintesi su S26?' vs 'Cos'è Now Brief su S26?' → MATCH (sintesi=summary=Now Brief, same Samsung AI feature)\n"
            "  'Wie bewertet man das S26?' vs 'Wie zufrieden sind die Nutzer mit dem S26?' → MATCH (both ask about user satisfaction/ratings)\n\n"
            "NO MATCH — same product/topic but different question:\n"
            "  'Is S26 waterproof?' vs 'Is S26 good for selfies?' → NO MATCH (different aspects)\n"
            "  'Can I use Siri AI on iPhone 17?' vs 'Is it safe to use AI on iPhone 17?' → NO MATCH (different sub-questions)\n"
            "  'What is Now Brief on S26?' vs 'What is special about S26?' → NO MATCH (specific feature vs general overview)\n"
            "  'When can I pre-order S26?' vs 'Where can I buy S26?' → NO MATCH (timing vs location, different intent)\n\n"
            "IMPORTANT RULES:\n"
            "- A Samsung Galaxy AI feature may be referred to by its official name (Now Brief, Now Nudge, Circle to Search, etc.) "
            "OR by a descriptive translation in another language (sintesi=brief/summary, riassunto=summary, résumé=summary). Match these.\n"
            "- Different phrasings or translations of the same question = MATCH.\n"
            "- Same topic but asking different things = NO MATCH.\n\n"
            "If no KL matches, use null. Each KL can be matched to at most one BL.\n\n"
            f"BL keywords:\n{bl_lines}\n\n"
            f"KL keywords:\n{kl_lines}\n\n"
            "Return JSON: "
            "{\"matches\": [{\"bl\": 0, \"kl\": 2}, {\"bl\": 1, \"kl\": null}, ...]}\n"
            "Include an entry for every BL keyword."
        )
        try:
            resp   = _llm_call_with_retry(
                client.chat.completions.create,
                model=LLM_MODEL, messages=[{"role": "user", "content": prompt}],
                temperature=0, response_format={"type": "json_object"})
            parsed = json.loads(resp.choices[0].message.content)
            results = []
            for item in parsed.get("matches", []):
                bl_idx = item.get("bl")
                kl_idx = item.get("kl")
                if bl_idx is None or not (0 <= bl_idx < len(batch_bl)):
                    continue
                if kl_idx is None or not (0 <= kl_idx < len(ref_keys)):
                    continue
                results.append((batch_bl[bl_idx], ref_keys[kl_idx]))
            return results
        except Exception as e:
            print(f"  LLM 오류 ({grp}): {e}")
            return []

    def process_group(grp, batches_for_grp):
        """그룹 내 배치를 순차 처리 — 배정된 KL을 다음 배치에서 제외"""
        total_bl  = sum(len(b[0]) for b in batches_for_grp)
        total_kl  = len(batches_for_grp[0][1]) if batches_for_grp else 0
        print(f"  [{grp}] BL {total_bl}개 × KL {total_kl}개  ({len(batches_for_grp)}배치)")
        used_kl  = set()
        used_bl_grp = set()
        grp_result = []
        for batch_num, (batch_bl, ref_keys) in enumerate(batches_for_grp):
            # 이미 배정된 KL 제외
            avail_keys = [rk for rk in ref_keys if rk not in used_kl]
            if not avail_keys:
                break
            avail_bl = [bn for bn in batch_bl if bn not in used_bl_grp]
            if not avail_bl:
                continue
            pairs = process_batch(grp, avail_bl, avail_keys)
            for bl_norm, ref_key in pairs:
                if ref_key not in used_kl and bl_norm not in used_bl_grp:
                    used_kl.add(ref_key)
                    used_bl_grp.add(bl_norm)
                    grp_result.append((bl_norm, ref_key))
            print(f"  [{grp}] 배치 {batch_num+1}/{len(batches_for_grp)} "
                  f"BL {len(avail_bl)}개 → 매칭 {len(pairs)}개")
        print(f"  [{grp}] 완료: 총 {len(grp_result)}개 매칭 / BL {total_bl}개 중")
        return grp_result

    all_raw   = []
    with ThreadPoolExecutor(max_workers=LLM_PARALLEL) as executor:
        futures = {
            executor.submit(process_group, grp, batches): grp
            for grp, batches in grp_batches.items()
        }
        for future in as_completed(futures):
            grp = futures[future]
            grp_pairs = future.result()
            all_raw.extend(grp_pairs)

    result = [(ref_key, bl_norm, 85.0, "LLM최종") for bl_norm, ref_key in all_raw]
    return result


def llm_final_verify_all(bl_fuzzy: dict, ref1: dict, ref1_pool: dict,
                          bl_pool: set, bl_lang: dict, bl_orig: dict, client,
                          confirm_count: dict = None, locked: set = None,
                          lock_after: int = 3, verify_iter: int = 0) -> int:
    """
    최종 검수: 정확일치(bl_exact) 제외 모든 매칭 쌍 LLM 검증
    confirm_count:  {bl_norm: 연속 통과 횟수} — 누적 관리 (호출자가 유지)
    locked:         연속 lock_after회 통과 → 다음 회차 skip
    의미 불일치 → BL/KL pool 반환 + confirm_count 초기화
    returns: 해제된 쌍 수
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from collections import Counter

    if confirm_count is None: confirm_count = {}
    if locked        is None: locked        = set()

    candidates = {}
    for bl_norm, (kl_text, score, label) in list(bl_fuzzy.items()):
        if bl_norm in locked:
            continue                        # 이미 lock_after회 연속 통과 → skip
        grp     = bl_lang.get(bl_norm, "")
        ref_key = (normalize(kl_text), grp)
        if ref_key in ref1:
            candidates[bl_norm] = (kl_text, score, label, ref_key)

    if not candidates:
        print(f"  검수 대상 없음  (locked skip: {len(locked):,}쌍)")
        return 0

    label_cnt = Counter(v[2] for v in candidates.values())
    print(f"  검수 대상: {len(candidates):,}쌍  locked skip: {len(locked):,}쌍  " +
          "  ".join(f"{lbl}:{cnt}" for lbl, cnt in sorted(label_cnt.items())))

    items   = list(candidates.items())
    batches = []
    for start in range(0, len(items), LLM_BATCH_SIZE):
        batch = items[start:start + LLM_BATCH_SIZE]
        pairs = [(bl_orig.get(bl_norm, bl_norm[0]), kl_text)
                 for bl_norm, (kl_text, _, _, _) in batch]
        batches.append((batch, pairs))

    rejected_bl = set()
    completed   = 0
    total       = len(batches)

    with ThreadPoolExecutor(max_workers=LLM_PARALLEL) as executor:
        futures = {executor.submit(ask_llm_verify_batch, client, pairs): batch
                   for batch, pairs in batches}
        for future in as_completed(futures):
            batch    = futures[future]
            verdicts = future.result()
            completed += 1
            print(f"  검수 배치 {completed}/{total} 완료")
            for (bl_norm, _), verdict in zip(batch, verdicts):
                if not verdict:
                    rejected_bl.add(bl_norm)

    newly_locked = 0
    for bl_norm in candidates:
        if bl_norm in rejected_bl:
            # 탈락 → count 초기화 + 조합 블랙리스트 등록
            confirm_count.pop(bl_norm, None)
            locked.discard(bl_norm)
            kl_text, score, label, ref_key = candidates[bl_norm]
            del bl_fuzzy[bl_norm]
            ref1_pool[ref_key] = kl_text
            bl_pool.add(bl_norm)
        else:
            # 통과 → count 증가
            confirm_count[bl_norm] = confirm_count.get(bl_norm, 0) + 1
            score = candidates[bl_norm][1]
            needed = 2 if score >= 90 else lock_after   # ≥90%는 2회, 미만은 lock_after회
            if confirm_count[bl_norm] >= needed:
                locked.add(bl_norm)
                newly_locked += 1
                # 레이블에 확정 차시 기록
                kl_text, score, label = bl_fuzzy[bl_norm]
                bl_fuzzy[bl_norm] = (kl_text, score, f"{label}(7-{verify_iter}회차)")

    kept = len(candidates) - len(rejected_bl)
    print(f"  검수 통과: {kept:,}쌍  오매칭 해제: {len(rejected_bl):,}쌍  "
          f"신규 locked(≥90%→2회/미만→{lock_after}회): {newly_locked:,}쌍")
    return len(rejected_bl)


def _fmt_elapsed(seconds: float) -> str:
    m, s = divmod(int(seconds), 60)
    return f"{m}분 {s}초" if m else f"{s}초"


def run(path_in: str, path_out: str):
    from openai import OpenAI
    client = OpenAI(api_key=OPENAI_API_KEY)

    run_start = time.perf_counter()
    print("키워드 로드 중...")
    ref1 = load_ref(path_in, REF_SHEET, REF_HEADER_ROW)
    grp_counts = defaultdict(int)
    for (_, grp) in ref1:
        grp_counts[grp] += 1
    print(f"  기존 키워드리스트: {len(ref1):,}개")
    for grp, cnt in sorted(grp_counts.items()):
        print(f"    {grp}: {cnt:,}개")

    wb = load_workbook(path_in)
    ws = wb[BL_SHEET]

    col_c = BL_RESULT_COL
    col_d = BL_RESULT_COL + 1
    col_e = BL_RESULT_COL + 2

    if BL_HEADER_ROW > 0:
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF")
        for col, name in [(col_c, "기존목록 매칭"), (col_d, "유사도"), (col_e, "배정차시")]:
            cell = ws.cell(BL_HEADER_ROW, col, name)
            cell.fill = hdr_fill; cell.font = hdr_font

    data_start = BL_HEADER_ROW + 1 if BL_HEADER_ROW > 0 else 1

    # ── BL 수집 ────────────────────────────────────────────────
    print("\n1단계: 정확 일치 처리 중...")
    bl_norm_to_rows = {}
    bl_lang         = {}
    bl_orig         = {}

    for row_idx in range(data_start, ws.max_row + 1):
        kw_val   = ws.cell(row_idx, BL_KW_COL).value
        lang_val = ws.cell(row_idx, BL_LANG_COL).value
        if not kw_val or str(kw_val).strip() == "":
            continue
        orig     = str(kw_val).strip()
        lang_grp = str(lang_val).strip().lower() if lang_val else ""
        kw_norm  = normalize(orig)
        bl_key   = (kw_norm, lang_grp)   # MX_de / Co.A_de 구분
        bl_norm_to_rows.setdefault(bl_key, []).append(row_idx)
        bl_lang[bl_key] = lang_grp
        bl_orig[bl_key] = orig

    exact_used   = set()
    bl_exact     = {}
    bl_unmatched = set()

    for bl_key in bl_norm_to_rows:
        kw_norm, grp = bl_key
        ref_key = (kw_norm, grp)
        if ref_key in ref1:
            bl_exact[bl_key] = (ref1[ref_key], 100, "정확일치")
            exact_used.add(ref_key)
        else:
            bl_unmatched.add(bl_key)

    ref1_pool = {k: v for k, v in ref1.items() if k not in exact_used}
    bl_pool   = set(bl_unmatched)
    bl_fuzzy  = {}
    iteration = 0

    print(f"  정확 일치: {len(bl_exact):,}개")
    print(f"  BL 불일치: {len(bl_pool):,}개  /  KL 불일치: {len(ref1_pool):,}개")

    # ── 일본어 브랜드 선행 매칭 ────────────────────────────────
    ja_bl_pre = sum(1 for bn in bl_pool if bl_lang.get(bn, "").endswith("_ja"))
    if ref1_pool and bl_pool and ja_bl_pre > 0:
        print(f"\n일본어 브랜드 선행 매칭... (_ja BL: {ja_bl_pre:,}개)")
        while ref1_pool and bl_pool:
            pairs = ja_brand_pairs(ref1_pool, bl_pool, bl_lang)
            if not pairs: break
            used_ref = set(); used_bl = set(); new_assigned = []
            for ref_key, bl_norm, score in pairs:
                if ref_key in used_ref or bl_norm in used_bl: continue
                used_ref.add(ref_key); used_bl.add(bl_norm)
                new_assigned.append((ref_key, bl_norm, score))
            if not new_assigned: break
            for ref_key, bl_norm, score in new_assigned:
                bl_fuzzy[bl_norm] = (ref1_pool[ref_key], score, "일본어브랜드")
                del ref1_pool[ref_key]; bl_pool.discard(bl_norm)
            print(f"  {len(new_assigned):,}개 배정  "
                  f"(남은 KL: {len(ref1_pool):,} / 남은 BL: {len(bl_pool):,})")

    # ── 2단계: 문자/단어 PHASES ────────────────────────────────
    for mode, threshold in PHASES:
        if not ref1_pool or not bl_pool: break
        iteration += 1
        mode_str = "문자" if mode == "char" else "단어"
        label = f"{iteration}차({mode_str})"
        print(f"\n{label} 유사도 계산 중... "
              f"(임계값 {threshold}%+  /  KL {len(ref1_pool):,} × BL {len(bl_pool):,})")
        pairs = calc_pairs(ref1_pool, bl_pool, bl_lang, mode=mode, min_score=threshold)
        if not pairs:
            print("  ※ 후보 없음 → 다음 단계")
            continue
        used_ref = set(); used_bl = set(); new_assigned = []
        for ref_key, bl_norm, score in pairs:
            if ref_key in used_ref or bl_norm in used_bl: continue
            used_ref.add(ref_key); used_bl.add(bl_norm)
            if score >= threshold: new_assigned.append((ref_key, bl_norm, score))
        if not new_assigned:
            print(f"  ※ {threshold}%+ 확정 쌍 없음 → 다음 단계")
            continue
        for ref_key, bl_norm, score in new_assigned:
            bl_fuzzy[bl_norm] = (ref1_pool[ref_key], score, label)
            del ref1_pool[ref_key]; bl_pool.discard(bl_norm)
        print(f"  {len(new_assigned):,}개 배정  "
              f"(남은 KL: {len(ref1_pool):,} / 남은 BL: {len(bl_pool):,})")

    # ── 일본어 브랜드 후행 매칭 ────────────────────────────────
    ja_bl_post = sum(1 for bn in bl_pool if bl_lang.get(bn, "").endswith("_ja"))
    if ref1_pool and bl_pool and ja_bl_post > 0:
        print(f"\n일본어 브랜드 후행 매칭... (_ja BL: {ja_bl_post:,}개)")
        while ref1_pool and bl_pool:
            pairs = ja_brand_pairs(ref1_pool, bl_pool, bl_lang)
            if not pairs: break
            used_ref = set(); used_bl = set(); new_assigned = []
            for ref_key, bl_norm, score in pairs:
                if ref_key in used_ref or bl_norm in used_bl: continue
                used_ref.add(ref_key); used_bl.add(bl_norm)
                new_assigned.append((ref_key, bl_norm, score))
            if not new_assigned: break
            for ref_key, bl_norm, score in new_assigned:
                bl_fuzzy[bl_norm] = (ref1_pool[ref_key], score, "일본어브랜드")
                del ref1_pool[ref_key]; bl_pool.discard(bl_norm)
            print(f"  {len(new_assigned):,}개 배정  "
                  f"(남은 KL: {len(ref1_pool):,} / 남은 BL: {len(bl_pool):,})")

    # ── 4단계: SBERT ≥80% ──────────────────────────────────────
    sbert_model = None
    if ref1_pool and bl_pool:
        print(f"\nSBERT 매칭 시작... "
              f"(임계값 {SBERT_MATCH_THRESHOLD}%+  /  "
              f"KL {len(ref1_pool):,} × BL {len(bl_pool):,})")
        sbert_pairs, sbert_model = sbert_match_all(ref1_pool, bl_pool, bl_lang, bl_orig)

        confirmed_direct = []
        to_verify        = {}

        for ref_key, bl_norm, score, label in sbert_pairs:
            if score >= SBERT_CONFIRM_THRESHOLD and label != "검토필요":
                confirmed_direct.append((ref_key, bl_norm, score, label))
            else:
                to_verify[bl_norm] = (ref_key, ref1_pool[ref_key], score, label)

        for ref_key, bl_norm, score, label in confirmed_direct:
            bl_fuzzy[bl_norm] = (ref1_pool[ref_key], score, label)
            del ref1_pool[ref_key]; bl_pool.discard(bl_norm)
        print(f"  SBERT ≥90% 확정: {len(confirmed_direct):,}개")

        if to_verify:
            for bl_norm, (ref_key, kl_text, score, label) in to_verify.items():
                if ref_key in ref1_pool: del ref1_pool[ref_key]
                bl_pool.discard(bl_norm)

            print(f"\nLLM 검증 시작... (80~89% 매칭 {len(to_verify):,}개)")
            confirmed_llm, rejected = llm_verify_pairs(client, to_verify, bl_orig, bl_lang)

            for bl_norm, (ref_key, kl_text, score, label) in confirmed_llm.items():
                bl_fuzzy[bl_norm] = (kl_text, score, label)
            print(f"  LLM 검증 통과: {len(confirmed_llm):,}개  /  "
                  f"미매칭 전환: {len(rejected):,}개")

            for bl_norm, (ref_key, kl_text) in rejected.items():
                bl_pool.add(bl_norm)
                ref1_pool[ref_key] = kl_text

        print(f"  SBERT 단계 완료  "
              f"(남은 KL: {len(ref1_pool):,} / 남은 BL: {len(bl_pool):,})")

    # ── 5단계: 미매칭 → LLM 매칭 (SBERT ≥50% 후보) ───────────
    if ref1_pool and bl_pool and sbert_model is not None:
        print(f"\nLLM 매칭 시작... "
              f"(SBERT 후보 임계값 {SBERT_CAND_THRESHOLD}%+  /  "
              f"KL {len(ref1_pool):,} × BL {len(bl_pool):,})")
        llm_pairs = llm_match(ref1_pool, bl_pool, bl_lang, bl_orig,
                              sbert_model, client)
        for ref_key, bl_norm, score, label in llm_pairs:
            bl_fuzzy[bl_norm] = (ref1_pool[ref_key], score, label)
            del ref1_pool[ref_key]; bl_pool.discard(bl_norm)
        print(f"  LLM 배정: {len(llm_pairs):,}개  "
              f"(남은 KL: {len(ref1_pool):,} / 남은 BL: {len(bl_pool):,})")

    # ── 6단계: 남은 KL 전체를 LLM에 직접 전송 (SBERT 없음) ───
    if ref1_pool and bl_pool:
        print(f"\nLLM 최종 패스 (SBERT 없이 남은 KL 전체 비교)... "
              f"KL {len(ref1_pool):,} × BL {len(bl_pool):,}")
        llm_pairs2 = llm_match_direct(ref1_pool, bl_pool, bl_lang, bl_orig, client)
        for ref_key, bl_norm, score, label in llm_pairs2:
            bl_fuzzy[bl_norm] = (ref1_pool[ref_key], score, label)
            del ref1_pool[ref_key]; bl_pool.discard(bl_norm)
        print(f"  LLM 최종 배정: {len(llm_pairs2):,}개  "
              f"(남은 KL: {len(ref1_pool):,} / 남은 BL: {len(bl_pool):,})")

    # ── 7단계: 최종 검수 반복 (3회 연속 통과 → locked skip) ─────
    MAX_VERIFY_ITER = 10
    LOCK_AFTER      = 3   # 연속 N회 통과 시 이후 회차 skip
    if bl_fuzzy:
        confirm_count  = {}   # {bl_norm: 연속 통과 횟수}
        locked         = set() # 3회 연속 통과 → skip 대상
        for verify_iter in range(1, MAX_VERIFY_ITER + 1):
            active = len(bl_fuzzy) - len(locked)
            print(f"\n7단계 {verify_iter}회차: 검수 대상 {active:,}쌍 "
                  f"(전체 {len(bl_fuzzy):,} - locked {len(locked):,})...")
            rejected_cnt = llm_final_verify_all(
                bl_fuzzy, ref1, ref1_pool, bl_pool, bl_lang, bl_orig, client,
                confirm_count=confirm_count, locked=locked, lock_after=LOCK_AFTER,
                verify_iter=verify_iter)

            if rejected_cnt == 0 and active == len(locked):
                print(f"  모든 쌍 통과 또는 locked → {verify_iter}회차 수렴")
                break

            if not (ref1_pool and bl_pool):
                print(f"  재매칭할 pool 없음 → 종료")
                break

            print(f"\n  재매칭 시작... KL {len(ref1_pool):,} × BL {len(bl_pool):,}")
            llm_pairs3 = llm_match_direct(ref1_pool, bl_pool, bl_lang, bl_orig, client)

            if not llm_pairs3:
                print(f"  새 매칭 없음 → 종료")
                break

            # 새 매칭 추가 → confirm_count 초기화 (새 쌍이므로 0부터)
            added = 0
            for ref_key, bl_norm, score, label in llm_pairs3:
                kl_text = ref1_pool[ref_key]
                bl_fuzzy[bl_norm] = (kl_text, score, "LLM재매칭")
                confirm_count.pop(bl_norm, None)
                locked.discard(bl_norm)
                del ref1_pool[ref_key]; bl_pool.discard(bl_norm)
                added += 1
            print(f"  재매칭 {added:,}개 추가 → 다음 회차 검수")
        else:
            print(f"  최대 {MAX_VERIFY_ITER}회 반복 완료")

        total_pairs   = len(bl_fuzzy)
        total_locked  = len(locked)
        total_verifs  = sum(min(c, MAX_VERIFY_ITER) for c in confirm_count.values())
        max_verifs    = total_pairs * MAX_VERIFY_ITER
        saved_verifs  = max_verifs - total_verifs
        print(f"\n[7단계 토큰 절약 현황]")
        print(f"  전체 매칭 쌍: {total_pairs:,}개  |  locked: {total_locked:,}개 "
              f"({total_locked/total_pairs*100:.0f}%)" if total_pairs else "")
        print(f"  실제 검증 호출: {total_verifs:,}회  |  최대치 대비 절약: {saved_verifs:,}회 "
              f"({saved_verifs/max_verifs*100:.0f}%)" if max_verifs else "")

        # 루프 종료 후 한 번도 검증 못 받은 재매칭 쌍 최종 검수
        unverified = [bn for bn in bl_fuzzy if confirm_count.get(bn, 0) == 0]
        if unverified:
            print(f"\n7단계 최종 보정 검수: 미검증 재매칭 {len(unverified):,}쌍...")
            llm_final_verify_all(
                bl_fuzzy, ref1, ref1_pool, bl_pool, bl_lang, bl_orig, client,
                confirm_count=confirm_count, locked=locked, lock_after=LOCK_AFTER,
                verify_iter=MAX_VERIFY_ITER + 1)

    # ── 최종 미매칭 현황 출력 ──────────────────────────────────
    if bl_pool or ref1_pool:
        # 그룹별로 미매칭 BL / KL 출력
        unmatched_bl_by_grp = defaultdict(list)
        for bn in bl_pool:
            unmatched_bl_by_grp[bl_lang.get(bn, "?")].append(bl_orig.get(bn, bn[0]))
        unmatched_kl_by_grp = defaultdict(list)
        for (kn, grp) in ref1_pool:
            unmatched_kl_by_grp[grp].append(ref1_pool[(kn, grp)])

        all_grps = sorted(set(list(unmatched_bl_by_grp) + list(unmatched_kl_by_grp)))
        print(f"\n[최종 미매칭 현황]")
        for grp in all_grps:
            bl_list = unmatched_bl_by_grp.get(grp, [])
            kl_list = unmatched_kl_by_grp.get(grp, [])
            if not bl_list and not kl_list:
                continue
            print(f"  {grp}: BL {len(bl_list)}개 미매칭 / KL {len(kl_list)}개 미매칭")
            for t in bl_list[:5]:
                print(f"    BL: {t}")
            if len(bl_list) > 5:
                print(f"    ... (+{len(bl_list)-5}개)")
            for t in kl_list[:5]:
                print(f"    KL: {t}")
            if len(kl_list) > 5:
                print(f"    ... (+{len(kl_list)-5}개)")

    # ── 결과 쓰기 ──────────────────────────────────────────────
    total = ws.max_row - data_start + 1
    cnt   = {"exact": 0, "fuzzy": 0, "none": 0}
    gray  = PatternFill("solid", fgColor="F2F2F2")
    print(f"\n결과 쓰기 중... (총 {total:,}행)")

    for row_idx in range(data_start, ws.max_row + 1):
        kw_val   = ws.cell(row_idx, BL_KW_COL).value
        lang_val = ws.cell(row_idx, BL_LANG_COL).value
        if not kw_val or str(kw_val).strip() == "":
            continue
        kw_norm  = normalize(str(kw_val).strip())
        lang_grp = str(lang_val).strip().lower() if lang_val else ""
        bl_key   = (kw_norm, lang_grp)

        if bl_key in bl_exact:
            c_val, c_score, c_label = bl_exact[bl_key]; cnt["exact"] += 1
        elif bl_key in bl_fuzzy:
            c_val, c_score, c_label = bl_fuzzy[bl_key]; cnt["fuzzy"] += 1
        else:
            c_val, c_score, c_label = "미매칭", 0, ""; cnt["none"] += 1

        if   c_label == "LLM매칭":    fill = PatternFill("solid", fgColor="B3FFB3")  # 연초록
        elif c_label == "LLM최종":    fill = PatternFill("solid", fgColor="FFE699")  # 연노랑
        elif c_label == "LLM재매칭":  fill = PatternFill("solid", fgColor="C8E6C9")  # 진연초록
        elif c_label == "LLM검증":    fill = PatternFill("solid", fgColor="B3D9FF")  # 연파랑
        elif c_label == "검토필요":   fill = PatternFill("solid", fgColor="FFB3DE")  # 연분홍
        elif c_label == "SBERT":      fill = PatternFill("solid", fgColor="D9B3FF")  # 연보라
        else:                          fill = row_fill(c_score) if c_score > 0 else gray

        ws.cell(row_idx, col_c, c_val).fill                                = fill
        ws.cell(row_idx, col_d, f"{c_score}%" if c_score > 0 else "").fill = fill
        ws.cell(row_idx, col_e, c_label).fill                              = fill

        if (row_idx - data_start + 1) % 3000 == 0:
            print(f"  {row_idx - data_start + 1:,}/{total:,}...")

    ws.column_dimensions[get_column_letter(col_c)].width = 70
    ws.column_dimensions[get_column_letter(col_d)].width = 16
    ws.column_dimensions[get_column_letter(col_e)].width = 12

    # ── KL 시트 복사 후 BL 매칭 열 추가 ─────────────────────────
    print("\nKL 시트에 BL 매칭 열 추가 중...")
    ws_kl         = wb[REF_SHEET]
    n_kl_cols     = max(max(cols) for cols in KL_LANG_COLS.values())
    data_start_kl = REF_HEADER_ROW + 1
    offset        = n_kl_cols  # BL 매칭 열 = KL 열 번호 + offset

    # KL 텍스트 → 원본 BL 텍스트 역방향 룩업 (언어 그룹 포함)
    kl_to_bl = {}
    for bl_norm, (kl_text, _, _) in bl_exact.items():
        grp = bl_lang.get(bl_norm, "")
        kl_to_bl[(normalize(kl_text), grp)] = bl_orig.get(bl_norm, bl_norm)
    for bl_norm, (kl_text, _, _) in bl_fuzzy.items():
        grp = bl_lang.get(bl_norm, "")
        kl_to_bl[(normalize(kl_text), grp)] = bl_orig.get(bl_norm, bl_norm)

    # 헤더 (원본 열 그대로, BL 매칭 열은 뒤에 추가)
    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    hdr_font  = Font(bold=True, color="FFFFFF")
    # BL매칭 텍스트 열 + 매핑여부 열 헤더
    for i in range(1, n_kl_cols + 1):
        kl_hdr = ws_kl.cell(REF_HEADER_ROW, i).value or ""
        bl_col  = i + offset
        chk_col = i + offset + n_kl_cols
        ws_kl.cell(REF_HEADER_ROW, bl_col,  f"{kl_hdr}_BL매칭").fill = hdr_fill
        ws_kl.cell(REF_HEADER_ROW, bl_col).font = hdr_font
        ws_kl.cell(REF_HEADER_ROW, chk_col, f"{kl_hdr}_매핑여부").fill = hdr_fill
        ws_kl.cell(REF_HEADER_ROW, chk_col).font = hdr_font

    bl_fill   = PatternFill("solid", fgColor="C8E6C9")   # 연초록 — 매핑됨
    miss_fill = PatternFill("solid", fgColor="FCE4D6")   # 연주황 — 미매핑
    hit_counts  = defaultdict(int)
    miss_counts = defaultdict(int)
    for row in range(data_start_kl, ws_kl.max_row + 1):
        for i in range(1, n_kl_cols + 1):
            kl_val = ws_kl.cell(row, i).value
            if not kl_val:
                continue
            grp      = _KL_COL_LANG.get(i - 1, "")
            kl_norm  = normalize(str(kl_val).strip())
            bl_match = kl_to_bl.get((kl_norm, grp), "")
            bl_col   = i + offset
            chk_col  = i + offset + n_kl_cols
            kl_ref  = f"{get_column_letter(i)}{row}"
            formula = f"=VLOOKUP({kl_ref},브랜드라이트!$C:$C,1,0)"
            if bl_match:
                ws_kl.cell(row, bl_col,  bl_match).fill = bl_fill
                ws_kl.cell(row, chk_col, formula)
                hit_counts[grp] += 1
            else:
                ws_kl.cell(row, chk_col, formula).fill = miss_fill
                miss_counts[grp] += 1

    print("  KL 룩업 결과 (그룹별 hit/miss):")
    all_grps = sorted(set(list(hit_counts) + list(miss_counts)))
    for grp in all_grps:
        h = hit_counts.get(grp, 0); m = miss_counts.get(grp, 0)
        print(f"    {grp}: hit={h}  miss={m}")

    wb.save(path_out)
    print()
    print("=" * 50)
    print(f"  정확 일치: {cnt['exact']:,}개")
    print(f"  유사 매칭: {cnt['fuzzy']:,}개")
    print(f"  미매칭:    {cnt['none']:,}개")
    print("=" * 50)
    print(f"저장 완료: {path_out}")
    print(f"총 소요시간: {_fmt_elapsed(time.perf_counter() - run_start)}")


def export_matched_keywords(path_out: str):
    """
    매칭된 BL 키워드만 추출 → 새 파일에 키워드 / 언어 / 브랜드 3열로 저장
    """
    from openpyxl import Workbook as WB

    wb_src = load_workbook(path_out, read_only=True, data_only=True)
    ws_src = wb_src[BL_SHEET]

    wb_out = WB()
    ws_out  = wb_out.active
    ws_out.title = "매핑 키워드"
    ws_miss = wb_out.create_sheet("미매칭 키워드")

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ws in (ws_out, ws_miss):
        for col, name in enumerate(["키워드", "언어", "브랜드"], start=1):
            c = ws.cell(1, col, name)
            c.fill = hdr_fill; c.font = hdr_font

    data_start = BL_HEADER_ROW + 1
    matched_cnt = 0; miss_cnt = 0
    for row in ws_src.iter_rows(min_row=data_start, values_only=True):
        kw      = row[BL_KW_COL - 1]
        lang    = row[BL_LANG_COL - 1]
        matched = row[BL_RESULT_COL - 1]
        if not kw:
            continue
        lang_str  = str(lang).strip() if lang else ""
        brand     = lang_str.rsplit("_", 1)[0] if "_" in lang_str else lang_str
        lang_code = lang_str.rsplit("_", 1)[1] if "_" in lang_str else lang_str
        if not matched or str(matched).strip() in ("", "미매칭"):
            ws_miss.append([str(kw).strip(), lang_code, brand])
            miss_cnt += 1
        else:
            ws_out.append([str(kw).strip(), lang_code, brand])
            matched_cnt += 1

    wb_src.close()

    export_path = path_out.replace(".xlsx", "_브랜드라이트 얼라인용 키워드리스트.xlsx")
    wb_out.save(export_path)
    print(f"매핑 키워드 추출 완료: 매칭 {matched_cnt:,}개 / 미매칭 {miss_cnt:,}개 → {export_path}")


if __name__ == "__main__":
    run(INPUT_FILE, OUTPUT_FILE)
    export_matched_keywords(OUTPUT_FILE)

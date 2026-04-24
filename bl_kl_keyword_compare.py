"""
브랜드라이트 & 키워드리스트 키워드 비교 자동화

[파일 구조]
키워드리스트 (KL): keyword, keyword_en, date, country, language, company, generic_branded, intent_lv1
브랜드라이트 (BL): keyword, keyword_en, language_code, company, generic_branded, intent_depth1

[처리 규칙]
- KL company 표준화: Samsung → MX, Apple → Co.A, Co.A → Co.A (원본 열 유지, 옆에 company_std 추가)
- KL language 소문자 정규화 (EN → en 등), 없으면 country로 파생
- BL 날짜: 파일명에서 추출 (예: _260301 → 2026-03-01)
- KL은 BL 날짜에 해당하는 행만 필터링
- 비교 키: company_std + language + keyword (모두 소문자)
- BL 파일 여러 개 지정 시 각각 비교 후 결과를 합쳐서 출력
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================================================
# 설정 영역
# ============================================================

# KL 파일 목록
KEYWORDLIST_FILES = [
    r"C:\Users\yunji\Documents\Samsung\맵핑 검수\S_core_initial_processing_v1.xlsx",
    r"C:\Users\yunji\OneDrive\문서\Buds\2026_MX GEO Miracle Gen AI Monitoring Prompts_all languages_260227_co.xlsx",
]

# 사용할 파일 인덱스 (0부터)
KEYWORDLIST_FILE_INDEX = 1

# 읽을 시트명
# - 비워두면 S_core 방식 (date 열 필터링)
# - 시트명 지정 시 열별 언어 방식 사용
KEYWORDLIST_SHEET = "Buds4 Prompts_300"

# 시트-열 방식일 때: 언어별 열 (알파벳)
KL_SHEET_LANG_COLS = {
    "en": "J",
    "ko": "K",
    "es": "L",
    "pt": "M",
    "de": "N",
    "it": "O",
    "fr": "P",
    "ja": "Q",
    "id": "R",
}

# 시트-열 방식일 때: company 고정값
KL_SHEET_COMPANY = "MX"

# BL 파일 1개: 문자열로
# BL 파일 여러 개: 리스트로
BRANDLIGHT_FILES = [
    r"C:\Users\yunji\OneDrive\문서\Buds\MiracleWeekly_S26_1000_260405_Samsung Buds.xlsx",
    r"C:\Users\yunji\OneDrive\문서\Buds\MiracleWeekly_S26_1000_260412_Samsung Buds.xlsx",
    r"C:\Users\yunji\OneDrive\문서\Buds\MiracleWeekly_S26_1000_260419_Samsung Buds.xlsx",
]

OUTPUT_DIR = r"C:\Users\yunji\OneDrive\문서\Buds"

# KL 필터 날짜 직접 지정 (예: "2026-04-05")
# 비워두면 첫 번째 BL 파일명에서 자동 추출
KL_FILTER_DATE = "2026-03-22"

# KL company 표준화 매핑
KL_COMPANY_MAP = {
    "Samsung": "MX",
    "Apple":   "Co.A",
    "Co.A":    "Co.A",
}

# MiracleWeekly 포맷 (query/country_code/line_of_business) → BL 포맷 변환용
# line_of_business → company 매핑
LOB_TO_COMPANY = {
    "Samsung Galaxy S26": "MX",
    "Samsung Galaxy S25": "MX",
    "Samsung Galaxy":     "MX",
    "Samsung Buds":       "MX",
    "Samsung":            "MX",
    "iPhone":             "Co.A",
    "AirPods":            "Co.A",
    "Apple":              "Co.A",
}

# KL country → language_code 매핑 (language 열이 비어있을 때 사용)
COUNTRY_TO_LANG = {
    "US": "en",
    "UK": "en",
    "IN": "en",
    "AU": "en",
    "AE": "en",
    "BR": "pt",
    "DE": "de",
    "ES": "es",
    "FR": "fr",
    "ID": "id",
    "IT": "it",
    "JP": "ja",
    "KR": "ko",
}

# Overview 고정 행 순서 (Company, Language)
OVERVIEW_ORDER = [
    ("MX",   "en"), ("MX",   "id"), ("MX",   "es"), ("MX",   "pt"), ("MX",   "de"),
    ("MX",   "it"), ("MX",   "ja"), ("MX",   "fr"), ("MX",   "ko"),
    ("Co.A", "en"), ("Co.A", "id"), ("Co.A", "es"), ("Co.A", "pt"), ("Co.A", "de"),
    ("Co.A", "it"), ("Co.A", "ja"), ("Co.A", "fr"), ("Co.A", "ko"),
]

# 정렬 우선순위 (있는 열만 적용)
KL_SORT_COLS = ["company_std", "language", "country", "keyword"]
BL_SORT_COLS = ["BL날짜", "company", "language_code", "keyword"]

# ============================================================


def extract_date_from_filename(filepath: str) -> str:
    """파일명 끝 6자리 숫자(YYMMDD)를 날짜 문자열(YYYY-MM-DD)로 변환"""
    stem = Path(filepath).stem
    match = re.search(r"(\d{6})$", stem)
    if not match:
        raise ValueError(f"파일명에서 날짜(YYMMDD)를 찾을 수 없습니다: {stem}")
    s = match.group(1)
    return f"20{s[0:2]}-{s[2:4]}-{s[4:6]}"


def extract_date_for_output(filepath: str) -> str:
    """출력 파일명용: 파일명 어디서든 6자리 숫자(YYMMDD) 추출"""
    stem = Path(filepath).stem
    match = re.search(r"(\d{6})", stem)
    if not match:
        raise ValueError(f"파일명에서 날짜(YYMMDD)를 찾을 수 없습니다: {stem}")
    s = match.group(1)
    return f"20{s[0:2]}-{s[2:4]}-{s[4:6]}"


def load_excel(filepath: str) -> pd.DataFrame:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {filepath}")
    return pd.read_excel(filepath)


def prepare_kl(kl_raw: pd.DataFrame, filter_date: str) -> pd.DataFrame:
    df = kl_raw.copy()

    # 날짜 표준화 및 필터링
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df[df["date"] == filter_date].copy()

    if df.empty:
        raise ValueError(f"KL에서 날짜 {filter_date}에 해당하는 데이터가 없습니다.\n"
                         f"KL에 있는 날짜: {sorted(pd.to_datetime(kl_raw['date'], errors='coerce').dt.strftime('%Y-%m-%d').dropna().unique().tolist())}")

    # language: 값이 있으면 소문자 정규화, NaN이면 country 매핑으로 파생
    lang = df["language"].astype(str).str.strip().str.lower()
    lang = lang.where(lang.notna() & (lang != "nan") & (lang != ""), None)
    country_derived = df["country"].astype(str).str.strip().map(COUNTRY_TO_LANG)
    df["language"] = lang.combine_first(country_derived)

    # company_std 열 추가 (company 바로 다음 위치)
    df["company_std"] = df["company"].astype(str).str.strip().map(KL_COMPANY_MAP).fillna(df["company"].astype(str).str.strip())
    cols = list(df.columns)
    cols.remove("company_std")
    cols.insert(cols.index("company") + 1, "company_std")
    df = df[cols]

    return df


def prepare_kl_from_sheets(filepath: str, sheets: list, lang_cols: dict,
                           company_std: str, filter_date: str) -> pd.DataFrame:
    """시트별 열에서 언어별 키워드를 읽어 KL DataFrame으로 변환"""
    dfs = []
    for sheet_name in sheets:
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=0)
        except Exception as e:
            print(f"  시트 '{sheet_name}' 로드 실패: {e}")
            continue
        for lang_code, col_letter in lang_cols.items():
            col_idx = ord(col_letter.upper()) - ord("A")
            if col_idx >= len(df.columns):
                continue
            keywords = df.iloc[:, col_idx].dropna().astype(str).str.strip()
            keywords = keywords[keywords != ""]
            if keywords.empty:
                continue
            dfs.append(pd.DataFrame({
                "keyword":     keywords.values,
                "language":    lang_code,
                "company_std": company_std,
                "date":        filter_date,
            }))
        print(f"  시트 '{sheet_name}' 로드 완료")
    if not dfs:
        return pd.DataFrame(columns=["keyword", "language", "company_std", "date"])
    return pd.concat(dfs, ignore_index=True)


def normalize_bl_format(df: pd.DataFrame) -> pd.DataFrame:
    """
    MiracleWeekly 신 포맷(query/country_code/line_of_business/date)을
    기존 BL 포맷(keyword/keyword_en/language_code/company/generic_branded/intent_depth1)으로 변환.
    이미 구 포맷이면 그대로 반환.
    """
    if "query" not in df.columns:
        return df  # 이미 구 포맷

    df = df.copy()
    df.rename(columns={"query": "keyword"}, inplace=True)

    # country_code → language_code
    df["language_code"] = (
        df["country_code"].astype(str).str.strip().str.upper().map(COUNTRY_TO_LANG)
    )

    # line_of_business → company (가장 긴 매핑 키부터 매칭)
    sorted_keys = sorted(LOB_TO_COMPANY.keys(), key=len, reverse=True)
    def map_lob(val):
        v = str(val).strip()
        for key in sorted_keys:
            if key.lower() in v.lower():
                return LOB_TO_COMPANY[key]
        return v
    df["company"] = df["line_of_business"].apply(map_lob)

    # 구 포맷에는 있지만 신 포맷에 없는 열은 빈 값으로 추가
    for col in ["keyword_en", "generic_branded", "intent_depth1"]:
        if col not in df.columns:
            df[col] = None

    # 불필요한 신 포맷 열 제거
    df.drop(columns=["country_code", "line_of_business", "date"], errors="ignore", inplace=True)

    return df[["keyword", "keyword_en", "language_code", "company", "generic_branded", "intent_depth1"]]


def prepare_bl(bl_raw: pd.DataFrame, bl_date: str) -> pd.DataFrame:
    df = normalize_bl_format(bl_raw)
    df["date"] = bl_date
    df["language_code"] = df["language_code"].astype(str).str.strip().str.lower()
    df["company"] = df["company"].astype(str).str.strip()
    df["keyword"] = df["keyword"].astype(str).str.strip()
    return df


def make_key(company: pd.Series, language: pd.Series, keyword: pd.Series) -> pd.Series:
    return company.str.lower() + "|" + language.str.lower() + "|" + keyword.str.lower()


def compare(kl: pd.DataFrame, bl: pd.DataFrame):
    kl["_key"] = make_key(kl["company_std"], kl["language"], kl["keyword"])
    bl["_key"] = make_key(bl["company"], bl["language_code"], bl["keyword"])

    kl_key_set = set(kl["_key"].dropna())
    bl_key_set = set(bl["_key"].dropna())

    both_keys    = kl_key_set & bl_key_set
    kl_only_keys = kl_key_set - bl_key_set
    bl_only_keys = bl_key_set - kl_key_set

    kl_only    = kl[kl["_key"].isin(kl_only_keys)].drop(columns=["_key"])
    bl_only    = bl[bl["_key"].isin(bl_only_keys)].drop(columns=["_key"])
    kl_matched = kl[kl["_key"].isin(both_keys)].drop(columns=["_key"])
    bl_matched = bl[bl["_key"].isin(both_keys)].drop(columns=["_key"])

    return kl_only, bl_only, kl_matched, bl_matched


def build_overview(kl_only, kl_matched, kl_date, bl_date):
    def count_group(df, company_col, lang_col, label):
        if df.empty:
            return pd.DataFrame(columns=["_company", "_language", label])
        g = df.groupby([company_col, lang_col]).size().reset_index(name=label)
        g.columns = ["_company", "_language", label]
        return g

    ov = pd.DataFrame(OVERVIEW_ORDER, columns=["_company", "_language"])
    ov = ov.merge(count_group(kl_only,    "company_std", "language", "키워드리스트 기준 브랜드라이트와 불일치"), on=["_company", "_language"], how="left")
    ov = ov.merge(count_group(kl_matched, "company_std", "language", "키워드리스트 기준 브랜드라이트와 일치"),   on=["_company", "_language"], how="left")
    ov = ov.fillna(0)
    for col in ["키워드리스트 기준 브랜드라이트와 불일치", "키워드리스트 기준 브랜드라이트와 일치"]:
        ov[col] = ov[col].astype(int)
    ov.insert(0, "BL날짜", bl_date)
    if kl_date is not None:
        ov.insert(0, "KL날짜", kl_date)
    ov.rename(columns={"_company": "Company", "_language": "Language"}, inplace=True)
    return ov


# ── Excel 출력 ──────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
KL_FILL     = PatternFill("solid", fgColor="D6E4F0")
BL_FILL     = PatternFill("solid", fgColor="D5F5E3")
ONLY_FILL   = PatternFill("solid", fgColor="FDEBD0")
DUPE_FONT   = Font(color="BFBFBF")


def auto_width(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame, row_fill=None,
                dupe_grey_cols: list = None, wrap_header: bool = False):
    ws = wb.create_sheet(title=name)
    display_cols = [c for c in df.columns if not c.startswith("_")]
    dupe_grey_cols = dupe_grey_cols or []

    for ci, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap_header)

    prev_vals = {}
    for ri, row in enumerate(df[display_cols].itertuples(index=False), 2):
        for ci, (col_name, value) in enumerate(zip(display_cols, row), 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            if row_fill:
                cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col_name in dupe_grey_cols and value == prev_vals.get(col_name):
                cell.font = DUPE_FONT
        for col_name, value in zip(display_cols, row):
            if col_name in dupe_grey_cols:
                prev_vals[col_name] = value

    auto_width(ws)
    return ws


def print_summary(overview_df: pd.DataFrame, kl_only_df: pd.DataFrame, bl_date: str):
    """Overview + KL에만 있는 키워드를 기반으로 요약 출력"""
    print(f"\n{'='*55}")
    print(f"  요약 ({bl_date})")
    print(f"{'='*55}")

    for company in overview_df["Company"].unique():
        rows = overview_df[overview_df["Company"] == company]
        if rows[["키워드리스트 기준 브랜드라이트와 불일치",
                  "키워드리스트 기준 브랜드라이트와 일치"]].sum().sum() == 0:
            continue
        print(f"\n{company}:")
        for _, row in rows.iterrows():
            lang     = row["Language"]
            matched  = int(row["키워드리스트 기준 브랜드라이트와 일치"])
            unmatched = int(row["키워드리스트 기준 브랜드라이트와 불일치"])
            total    = matched + unmatched
            if total == 0:
                continue
            line = f"  {lang.upper()}: {total}개 중 {matched}개 일치"
            if not kl_only_df.empty:
                miss_kws = kl_only_df[
                    (kl_only_df["company_std"] == company) &
                    (kl_only_df["language"] == lang)
                ]["keyword"].tolist()
                if miss_kws:
                    line += f"  → 불일치: {', '.join(miss_kws)}"
            print(line)


def sort_df(df: pd.DataFrame, sort_cols: list) -> pd.DataFrame:
    cols = [c for c in sort_cols if c in df.columns]
    if cols:
        df = df.sort_values(cols, ascending=True, ignore_index=True)
    return df


def add_bl_date_col(df: pd.DataFrame, bl_date: str) -> pd.DataFrame:
    """BL날짜 열을 맨 앞에 추가"""
    df = df.copy()
    df.insert(0, "BL날짜", bl_date)
    return df


def export_excel(output_path: str,
                 kl: pd.DataFrame,
                 all_overview: list, all_bl: list,
                 all_kl_only: list):

    wb = Workbook()
    wb.remove(wb.active)

    overview_df = pd.concat(all_overview, ignore_index=True)
    write_sheet(wb, "Overview", overview_df,
                dupe_grey_cols=["KL날짜", "BL날짜", "Company", "Language"],
                wrap_header=True)

    bl_all = pd.concat(all_bl, ignore_index=True)
    write_sheet(wb, "브랜드라이트", sort_df(bl_all, BL_SORT_COLS), BL_FILL)

    write_sheet(wb, "키워드리스트", sort_df(kl, KL_SORT_COLS), KL_FILL)

    kl_only_all = pd.concat(all_kl_only, ignore_index=True) if all_kl_only else pd.DataFrame()
    write_sheet(wb, "KL에만 있는 키워드", sort_df(kl_only_all, ["BL날짜"] + KL_SORT_COLS), ONLY_FILL)

    wb.save(output_path)
    print(f"\n저장 완료: {output_path}")


# ── main ────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  키워드 비교 자동화")
    print("=" * 55)

    # BL 파일 목록 정규화
    bl_files = BRANDLIGHT_FILES if isinstance(BRANDLIGHT_FILES, list) else [BRANDLIGHT_FILES]

    # KL 날짜: 직접 지정하면 그대로 사용, 아니면 첫 번째 BL 파일명에서 추출
    if KL_FILTER_DATE:
        kl_date = KL_FILTER_DATE
    else:
        kl_date = extract_date_from_filename(bl_files[0])

    # KL 로드 및 정돈 (한 번만)
    print(f"\n[1] 키워드리스트 로드")
    kl_file = KEYWORDLIST_FILES[KEYWORDLIST_FILE_INDEX]

    if KEYWORDLIST_SHEET:
        print(f"  파일[{KEYWORDLIST_FILE_INDEX}] 시트 '{KEYWORDLIST_SHEET}' - 열별 언어 방식")
        kl = prepare_kl_from_sheets(kl_file, [KEYWORDLIST_SHEET], KL_SHEET_LANG_COLS,
                                    KL_SHEET_COMPANY, kl_date)
    else:
        print(f"  파일[{KEYWORDLIST_FILE_INDEX}] S_core 방식 → {kl_date} 필터링")
        kl_raw = load_excel(kl_file)
        kl = prepare_kl(kl_raw, kl_date)

    print(f"  로드 완료: {len(kl):,}행  |  company_std: {sorted(kl['company_std'].unique().tolist())}")

    print(f"\n[2] 브랜드라이트 비교 및 저장 ({len(bl_files)}개 파일)")
    for bl_file in bl_files:
        bl_date = extract_date_for_output(bl_file)
        bl_raw = load_excel(bl_file)
        bl = prepare_bl(bl_raw, bl_date)

        kl_only, _, kl_matched, _ = compare(kl.copy(), bl.copy())

        print(f"\n  [{bl_date}]  KL만:{len(kl_only):,}  KL일치:{len(kl_matched):,}")

        output_date = extract_date_for_output(bl_file).replace("-", "")
        output_path = str(Path(OUTPUT_DIR) / f"키워드 비교_{output_date}.xlsx")
        overview = build_overview(kl_only, kl_matched, None if KEYWORDLIST_SHEET else kl_date, bl_date)
        export_excel(output_path, kl,
                     [overview],
                     [add_bl_date_col(bl, bl_date)],
                     [add_bl_date_col(kl_only, bl_date)])
        print_summary(overview, kl_only, bl_date)


if __name__ == "__main__":
    main()

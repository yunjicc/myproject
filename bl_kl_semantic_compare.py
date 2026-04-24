"""
브랜드라이트 & 키워드리스트 시맨틱 유사도 비교
fastembed 기반 (PyTorch 불필요, ONNX Runtime 사용)

[설치]
pip install fastembed scikit-learn openpyxl pandas

[처리 방식]
- KL / BL 키워드를 (company, language) 그룹별로 임베딩
- 코사인 유사도로 각 BL 키워드와 가장 가까운 KL 키워드를 매칭
- SIMILARITY_THRESHOLD 이상이면 일치로 판단
"""

import re
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import sys
from unittest.mock import MagicMock
for _mod in ["mmh3", "py_rust_stemmers"]:
    if _mod not in sys.modules:
        sys.modules[_mod] = MagicMock()

from fastembed.text import TextEmbedding
from sklearn.metrics.pairwise import cosine_similarity

# ============================================================
# 설정 영역
# ============================================================

KEYWORDLIST_FILES = [
    r"C:\Users\yunji\Documents\Samsung\맵핑 검수\S_core_initial_processing_v1.xlsx",
    r"C:\Users\yunji\OneDrive\문서\Buds\2026_MX GEO Miracle Gen AI Monitoring Prompts_all languages_260227_co.xlsx",
]
KEYWORDLIST_FILE_INDEX = 1
KEYWORDLIST_SHEET = "Miracle Prompts_1000 (2)"   # 비워두면 S_core 방식 (date 필터)

KL_SHEET_LANG_COLS = {
    "en": "J", "ko": "K", "es": "L", "pt": "M", "de": "N",
    "it": "O", "fr": "P", "ja": "Q", "id": "R",
}
KL_SHEET_COMPANY = "MX"

BRANDLIGHT_FILES = [
    r"C:\Users\yunji\OneDrive\문서\S26\MiracleWeekly_S26_1000_260308.xlsx",
    # r"C:\Users\yunji\OneDrive\문서\Buds\MiracleWeekly_S26_1000_260412_Samsung Buds.xlsx",
    # r"C:\Users\yunji\OneDrive\문서\Buds\MiracleWeekly_S26_1000_260419_Samsung Buds.xlsx",
]

OUTPUT_DIR    = r"C:\Users\yunji\OneDrive\문서\Buds"
KL_FILTER_DATE = "2026-03-22"   # S_core 방식일 때만 사용

# ── 시맨틱 설정 ──────────────────────────────────────────────
MODEL_NAME           = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"
SIMILARITY_THRESHOLD = 0.80   # 0~1, 이 이상이면 일치로 판단
BATCH_SIZE           = 128

# ── 매핑 테이블 ──────────────────────────────────────────────
KL_COMPANY_MAP = {"Samsung": "MX", "Apple": "Co.A", "Co.A": "Co.A"}

LOB_TO_COMPANY = {
    "Samsung Galaxy S26": "MX", "Samsung Galaxy S25": "MX",
    "Samsung Galaxy": "MX",     "Samsung Buds": "MX",
    "Samsung": "MX",            "iPhone": "Co.A",
    "AirPods": "Co.A",          "Apple": "Co.A",
}

COUNTRY_TO_LANG = {
    "US": "en", "UK": "en", "IN": "en", "AU": "en", "AE": "en",
    "BR": "pt", "DE": "de", "ES": "es", "FR": "fr",
    "ID": "id", "IT": "it", "JP": "ja", "KR": "ko",
}

OVERVIEW_ORDER = [
    ("MX",   "en"), ("MX",   "id"), ("MX",   "es"), ("MX",   "pt"), ("MX",   "de"),
    ("MX",   "it"), ("MX",   "ja"), ("MX",   "fr"), ("MX",   "ko"),
    ("Co.A", "en"), ("Co.A", "id"), ("Co.A", "es"), ("Co.A", "pt"), ("Co.A", "de"),
    ("Co.A", "it"), ("Co.A", "ja"), ("Co.A", "fr"), ("Co.A", "ko"),
]

# ============================================================


def extract_date_for_output(filepath: str) -> str:
    stem = Path(filepath).stem
    m = re.search(r"(\d{6})", stem)
    if not m:
        raise ValueError(f"파일명에서 날짜를 찾을 수 없습니다: {stem}")
    s = m.group(1)
    return f"20{s[0:2]}-{s[2:4]}-{s[4:6]}"


def load_excel(filepath: str) -> pd.DataFrame:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {filepath}")
    return pd.read_excel(filepath)


def prepare_kl(kl_raw: pd.DataFrame, filter_date: str) -> pd.DataFrame:
    df = kl_raw.copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df[df["date"] == filter_date].copy()
    if df.empty:
        raise ValueError(
            f"KL에서 날짜 {filter_date}에 해당하는 데이터가 없습니다.\n"
            f"KL에 있는 날짜: {sorted(pd.to_datetime(kl_raw['date'], errors='coerce').dt.strftime('%Y-%m-%d').dropna().unique().tolist())}"
        )
    lang = df["language"].astype(str).str.strip().str.lower()
    lang = lang.where(lang.notna() & (lang != "nan") & (lang != ""), None)
    df["language"] = lang.combine_first(df["country"].astype(str).str.strip().map(COUNTRY_TO_LANG))
    df["company_std"] = (
        df["company"].astype(str).str.strip()
        .map(KL_COMPANY_MAP)
        .fillna(df["company"].astype(str).str.strip())
    )
    return df


def prepare_kl_from_sheets(filepath: str, sheet: str, lang_cols: dict, company_std: str) -> pd.DataFrame:
    df = pd.read_excel(filepath, sheet_name=sheet, header=0)
    dfs = []
    for lang_code, col_letter in lang_cols.items():
        col_idx = ord(col_letter.upper()) - ord("A")
        if col_idx >= len(df.columns):
            continue
        keywords = df.iloc[:, col_idx].dropna().astype(str).str.strip()
        keywords = keywords[keywords != ""]
        if keywords.empty:
            continue
        dfs.append(pd.DataFrame({
            "keyword":     keywords.values,
            "language":    lang_code,
            "company_std": company_std,
        }))
    if not dfs:
        return pd.DataFrame(columns=["keyword", "language", "company_std"])
    return pd.concat(dfs, ignore_index=True)


def normalize_bl_format(df: pd.DataFrame) -> pd.DataFrame:
    if "query" not in df.columns:
        return df
    df = df.copy()
    df.rename(columns={"query": "keyword"}, inplace=True)
    df["language_code"] = df["country_code"].astype(str).str.strip().str.upper().map(COUNTRY_TO_LANG)
    sorted_keys = sorted(LOB_TO_COMPANY.keys(), key=len, reverse=True)
    def map_lob(val):
        v = str(val).strip()
        for key in sorted_keys:
            if key.lower() in v.lower():
                return LOB_TO_COMPANY[key]
        return v
    df["company"] = df["line_of_business"].apply(map_lob)
    for col in ["keyword_en", "generic_branded", "intent_depth1"]:
        if col not in df.columns:
            df[col] = None
    df.drop(columns=["country_code", "line_of_business", "date"], errors="ignore", inplace=True)
    return df[["keyword", "keyword_en", "language_code", "company", "generic_branded", "intent_depth1"]]


def prepare_bl(bl_raw: pd.DataFrame, bl_date: str) -> pd.DataFrame:
    df = normalize_bl_format(bl_raw)
    df["date"] = bl_date
    df["language_code"] = df["language_code"].astype(str).str.strip().str.lower()
    df["company"]       = df["company"].astype(str).str.strip()
    df["keyword"]       = df["keyword"].astype(str).str.strip()
    return df


# ── 시맨틱 비교 ──────────────────────────────────────────────

def semantic_compare(kl_df: pd.DataFrame, bl_df: pd.DataFrame,
                     model: TextEmbedding, threshold: float,
                     bl_date: str):
    """
    (company, language) 그룹별로 시맨틱 매칭 수행.

    Returns
    -------
    matched_df  : BL키워드 | KL키워드 | 유사도 | company | language | BL날짜
    kl_only_df  : 매칭되지 않은 KL 키워드
    """
    groups = sorted(set(zip(kl_df["company_std"], kl_df["language"])))

    matched_rows = []
    kl_only_rows = []

    for company, lang in groups:
        kl_grp = kl_df[
            (kl_df["company_std"] == company) & (kl_df["language"] == lang)
        ].reset_index(drop=True)

        bl_grp = bl_df[
            (bl_df["company"] == company) & (bl_df["language_code"] == lang)
        ].reset_index(drop=True)

        if bl_grp.empty:
            for kw in kl_grp["keyword"]:
                kl_only_rows.append({"BL날짜": bl_date, "company_std": company, "language": lang, "keyword": kw})
            continue

        kl_kws = kl_grp["keyword"].str.lower().tolist()
        bl_kws = bl_grp["keyword"].str.lower().tolist()

        print(f"    {company}/{lang}: KL {len(kl_kws)}개 × BL {len(bl_kws)}개")

        kl_emb = np.array(list(model.embed(kl_kws, batch_size=BATCH_SIZE)))
        bl_emb = np.array(list(model.embed(bl_kws, batch_size=BATCH_SIZE)))

        # sim[i, j] = BL[i] ↔ KL[j] 유사도
        sim = cosine_similarity(bl_emb, kl_emb)

        kl_matched = set()

        for bi in range(len(bl_kws)):
            best_j = int(np.argmax(sim[bi]))
            score  = float(sim[bi][best_j])
            if score >= threshold:
                matched_rows.append({
                    "BL날짜":  bl_date,
                    "company": company,
                    "language": lang,
                    "BL키워드": bl_grp.loc[bi, "keyword"],
                    "KL키워드": kl_grp.loc[best_j, "keyword"],
                    "유사도":   round(score, 4),
                })
                kl_matched.add(best_j)

        for ki in range(len(kl_kws)):
            if ki not in kl_matched:
                kl_only_rows.append({
                    "BL날짜":      bl_date,
                    "company_std": company,
                    "language":    lang,
                    "keyword":     kl_grp.loc[ki, "keyword"],
                })

    return pd.DataFrame(matched_rows), pd.DataFrame(kl_only_rows)


# ── Overview 집계 ────────────────────────────────────────────

def build_overview(matched_df: pd.DataFrame, kl_only_df: pd.DataFrame,
                   bl_date: str, kl_date):
    def count_grp(df, company_col, lang_col, label):
        if df.empty:
            return pd.DataFrame(columns=["_company", "_language", label])
        g = df.groupby([company_col, lang_col]).size().reset_index(name=label)
        g.columns = ["_company", "_language", label]
        return g

    ov = pd.DataFrame(OVERVIEW_ORDER, columns=["_company", "_language"])
    ov = ov.merge(
        count_grp(kl_only_df,  "company_std", "language", "키워드리스트 기준 브랜드라이트와 불일치"),
        on=["_company", "_language"], how="left"
    )
    ov = ov.merge(
        count_grp(matched_df,  "company",     "language", "키워드리스트 기준 브랜드라이트와 일치"),
        on=["_company", "_language"], how="left"
    )
    ov = ov.fillna(0)
    for col in ["키워드리스트 기준 브랜드라이트와 불일치", "키워드리스트 기준 브랜드라이트와 일치"]:
        ov[col] = ov[col].astype(int)
    ov.insert(0, "BL날짜", bl_date)
    if kl_date is not None:
        ov.insert(0, "KL날짜", kl_date)
    ov.rename(columns={"_company": "Company", "_language": "Language"}, inplace=True)
    return ov


# ── Excel 출력 ──────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
KL_FILL     = PatternFill("solid", fgColor="D6E4F0")
BL_FILL     = PatternFill("solid", fgColor="D5F5E3")
MATCH_FILL  = PatternFill("solid", fgColor="EAF4FB")
ONLY_FILL   = PatternFill("solid", fgColor="FDEBD0")
DUPE_FONT   = Font(color="BFBFBF")


def auto_width(ws):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame,
                row_fill=None, dupe_grey_cols: list = None, wrap_header: bool = False):
    ws = wb.create_sheet(title=name)
    display_cols  = [c for c in df.columns if not c.startswith("_")]
    dupe_grey_cols = dupe_grey_cols or []

    for ci, col_name in enumerate(display_cols, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap_header)

    prev_vals = {}
    for ri, row in enumerate(df[display_cols].itertuples(index=False), 2):
        for ci, (col_name, value) in enumerate(zip(display_cols, row), 1):
            cell = ws.cell(row=ri, column=ci, value=value)
            if row_fill:
                cell.fill = row_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col_name in dupe_grey_cols and value == prev_vals.get(col_name):
                cell.font = DUPE_FONT
        for col_name, value in zip(display_cols, row):
            if col_name in dupe_grey_cols:
                prev_vals[col_name] = value

    auto_width(ws)
    return ws


def sort_df(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    valid = [c for c in cols if c in df.columns]
    return df.sort_values(valid, ignore_index=True) if valid else df


def export_excel(output_path: str, kl_df: pd.DataFrame, bl_df: pd.DataFrame,
                 overview_df: pd.DataFrame, matched_df: pd.DataFrame,
                 kl_only_df: pd.DataFrame):
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(wb, "Overview", overview_df,
                dupe_grey_cols=["KL날짜", "BL날짜", "Company", "Language"],
                wrap_header=True)

    write_sheet(wb, "브랜드라이트", sort_df(bl_df,  ["company", "language_code", "keyword"]), BL_FILL)
    write_sheet(wb, "키워드리스트", sort_df(kl_df,  ["company_std", "language", "keyword"]),  KL_FILL)

    write_sheet(wb, "매칭 결과",
                sort_df(matched_df, ["BL날짜", "company", "language", "유사도"]),
                MATCH_FILL)

    write_sheet(wb, "KL에만 있는 키워드",
                sort_df(kl_only_df, ["BL날짜", "company_std", "language", "keyword"]),
                ONLY_FILL)

    wb.save(output_path)
    print(f"\n저장 완료: {output_path}")


# ── main ────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  키워드 시맨틱 비교 자동화")
    print(f"  모델: {MODEL_NAME}  |  임계값: {SIMILARITY_THRESHOLD}")
    print("=" * 55)

    bl_files = BRANDLIGHT_FILES if isinstance(BRANDLIGHT_FILES, list) else [BRANDLIGHT_FILES]

    # KL 날짜 (S_core 방식 전용)
    kl_date = KL_FILTER_DATE if KL_FILTER_DATE else None

    # KL 로드
    print(f"\n[1] 키워드리스트 로드")
    kl_file = KEYWORDLIST_FILES[KEYWORDLIST_FILE_INDEX]
    if KEYWORDLIST_SHEET:
        print(f"  파일[{KEYWORDLIST_FILE_INDEX}] 시트 '{KEYWORDLIST_SHEET}' - 열별 언어 방식")
        kl = prepare_kl_from_sheets(kl_file, KEYWORDLIST_SHEET, KL_SHEET_LANG_COLS, KL_SHEET_COMPANY)
        kl_date_label = None
    else:
        print(f"  파일[{KEYWORDLIST_FILE_INDEX}] S_core 방식 → {kl_date} 필터링")
        kl = prepare_kl(load_excel(kl_file), kl_date)
        kl_date_label = kl_date
    print(f"  로드 완료: {len(kl):,}행")

    # 모델 로드 (최초 1회)
    print(f"\n[2] 모델 로드: {MODEL_NAME}")
    model = TextEmbedding(MODEL_NAME)
    print("  로드 완료")

    # BL 파일별 처리
    print(f"\n[3] 브랜드라이트 비교 ({len(bl_files)}개 파일)")
    for bl_file in bl_files:
        bl_date = extract_date_for_output(bl_file)
        bl = prepare_bl(load_excel(bl_file), bl_date)
        print(f"\n  [{bl_date}] BL {len(bl):,}행")

        matched_df, kl_only_df = semantic_compare(kl, bl, model, SIMILARITY_THRESHOLD, bl_date)

        print(f"  → 매칭: {len(matched_df):,}  KL만: {len(kl_only_df):,}")

        overview = build_overview(matched_df, kl_only_df, bl_date, kl_date_label)

        output_date = bl_date.replace("-", "")
        output_path = str(Path(OUTPUT_DIR) / f"시맨틱비교_{output_date}.xlsx")
        export_excel(output_path, kl, bl, overview, matched_df, kl_only_df)


if __name__ == "__main__":
    main()
